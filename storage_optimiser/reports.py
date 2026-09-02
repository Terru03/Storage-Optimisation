import csv
from collections.abc import Iterable
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .analysis import (
    drive_summary,
    duplicate_group_summary,
    enrich_files,
    extension_summary,
    folder_summary,
    google_takeout_rows,
    human_bytes,
    possible_duplicate_folders,
)
from .cache import StorageCache
from .config import load_config
from .paths import ensure_app_dirs


HEADERS = {
    "path": "Full path",
    "folder": "Folder",
    "drive": "Drive",
    "filename": "Filename",
    "extension": "Extension",
    "size": "Size bytes",
    "size_human": "Size",
    "modified_at": "Modified date",
    "age_days": "Age days",
    "partial_hash": "Partial hash",
    "full_hash": "Full SHA-256 hash",
    "group_id": "Duplicate group",
    "duplicate_status": "Duplicate status",
    "duplicate_exists_on_b": "Duplicate exists on B:",
    "exact_duplicate_exists_on_b": "Exact duplicate exists on B:",
    "duplicate_only_outside_b": "Duplicate only outside B:",
    "same_filename_different_content": "Same filename but different content",
    "same_size_different_content": "Same size but different content",
    "only_found_on_p": "Only found on P:",
    "possible_old_editing_file": "Possible old editing file",
    "large_file": "Large file",
    "hardlinked_file": "Hardlinked file",
    "safe_to_review_manually": "Safe to review manually",
    "labels": "Review labels",
    "member_drives": "Matched drives",
    "folder_url": "Open folder link",
    "file_url": "Open file link",
    "phase": "Scan phase",
}


FILE_COLUMNS = [
    "group_id",
    "drive",
    "filename",
    "extension",
    "size",
    "size_human",
    "modified_at",
    "age_days",
    "duplicate_status",
    "exact_duplicate_exists_on_b",
    "duplicate_only_outside_b",
    "same_filename_different_content",
    "same_size_different_content",
    "only_found_on_p",
    "possible_old_editing_file",
    "large_file",
    "hardlinked_file",
    "safe_to_review_manually",
    "labels",
    "member_drives",
    "folder",
    "path",
    "folder_url",
    "file_url",
    "partial_hash",
    "full_hash",
]


def collect_report_data(
    config_path: str | Path | None = None,
    scan_id: int | None = None,
) -> dict[str, Any]:
    ensure_app_dirs()
    config = load_config(config_path)
    max_rows = min(50_000, max(1, int(config.get("report_max_rows_per_sheet", 200000))))
    cache = StorageCache(config["database_path"])
    try:
        if scan_id is None:
            scan_id = cache.latest_scan_id()
        if scan_id is None:
            raise RuntimeError("No complete scan found. Run scan first.")
        scan = cache.scan_row(scan_id)
        if scan is None:
            raise RuntimeError(f"Scan {scan_id} not found.")
        rows = cache.file_rows_filtered(scan_id, limit=max_rows)
        errors = cache.error_rows(scan_id, limit=max_rows)
    finally:
        cache.close()

    enriched, metrics = enrich_files(rows, config)
    duplicate_rows = [row for row in enriched if row.get("duplicate_status") == "Duplicate"]
    duplicate_groups = duplicate_group_summary(enriched)
    p_on_b = [
        row for row in enriched
        if row.get("drive") == metrics["portable_drive"] and row.get("exact_duplicate_exists_on_b")
    ]
    p_not_on_b = [
        row for row in enriched
        if row.get("drive") == metrics["portable_drive"] and not row.get("exact_duplicate_exists_on_b")
    ]
    not_on_b = [
        row for row in enriched
        if row.get("drive") != metrics["backup_drive"] and not row.get("exact_duplicate_exists_on_b")
    ]
    large_files = [row for row in enriched if row.get("large_file")]
    old_files = [row for row in enriched if int(row.get("age_days") or 0) >= metrics["old_file_days"]]
    top_files = sorted(enriched, key=lambda row: int(row.get("size") or 0), reverse=True)[:100]

    return {
        "config": config,
        "scan": scan,
        "scan_id": scan_id,
        "errors": errors,
        "metrics": metrics,
        "max_rows": max_rows,
        "tables": {
            "Duplicate Groups": duplicate_rows[:max_rows],
            "Duplicate Group Summary": duplicate_groups[:max_rows],
            "Files on P also found on B": p_on_b[:max_rows],
            "Files on P not found on B": p_not_on_b[:max_rows],
            "Files not found on B": not_on_b[:max_rows],
            "Top 100 Biggest Files": top_files,
            "Top 50 Biggest Folders": folder_summary(enriched)[:50],
            "Possible Duplicate Folders": possible_duplicate_folders(enriched)[:max_rows],
            "Google Takeout Review": google_takeout_rows(enriched)[:max_rows],
            "Large Files": large_files[:max_rows],
            "Old Files": old_files[:max_rows],
            "Extension Summary": extension_summary(enriched),
            "Drive Summary": drive_summary(enriched),
            "Folder Size Summary": folder_summary(enriched)[:max_rows],
            "Scan Errors": errors,
        },
    }


def export_excel_report(
    config_path: str | Path | None = None,
    output_path: str | Path | None = None,
    scan_id: int | None = None,
) -> Path:
    report = collect_report_data(config_path, scan_id)
    config = report["config"]
    scan = report["scan"]
    scan_id = int(report["scan_id"])
    metrics = report["metrics"]
    errors = report["errors"]
    max_rows = report["max_rows"]
    tables = report["tables"]

    wb = Workbook()
    wb.remove(wb.active)
    write_summary(wb, scan, metrics, len(errors), config)
    write_table(wb, "Duplicate Groups", tables["Duplicate Groups"], FILE_COLUMNS)
    write_table(
        wb,
        "Duplicate Group Summary",
        tables["Duplicate Group Summary"],
        [
            "group_id",
            "file_count",
            "file_size_bytes",
            "file_size",
            "wasted_bytes",
            "wasted_space",
            "involved_bytes",
            "involved_space",
            "drives",
            "sample_filename",
            "folders",
            "full_hash",
            "safe_to_review_manually",
        ],
        header_map={
            "group_id": "Duplicate group",
            "file_count": "File count",
            "file_size_bytes": "Each file bytes",
            "file_size": "Each file size",
            "wasted_bytes": "Review bytes",
            "wasted_space": "Review space",
            "involved_bytes": "Involved bytes",
            "involved_space": "Involved space",
            "drives": "Drives",
            "sample_filename": "Sample filename",
            "folders": "Folders",
            "full_hash": "Full SHA-256 hash",
            "safe_to_review_manually": "Safe to review manually",
        },
    )
    write_table(wb, "Files on P also found on B", tables["Files on P also found on B"], FILE_COLUMNS)
    write_table(wb, "Files on P not found on B", tables["Files on P not found on B"], FILE_COLUMNS)
    write_table(wb, "Files not found on B", tables["Files not found on B"], FILE_COLUMNS)
    write_table(wb, "Top 100 Biggest Files", tables["Top 100 Biggest Files"], FILE_COLUMNS)
    write_table(
        wb,
        "Top 50 Biggest Folders",
        tables["Top 50 Biggest Folders"],
        ["folder", "open_folder", "file_count", "total_bytes", "total_size"],
        header_map={
            "folder": "Folder",
            "open_folder": "Open folder link",
            "file_count": "File count",
            "total_bytes": "Total bytes",
            "total_size": "Total size",
        },
    )
    write_table(
        wb,
        "Possible Duplicate Folders",
        tables["Possible Duplicate Folders"],
        ["signature", "folder", "open_folder", "file_count", "total_bytes", "total_size", "matching_folder_count"],
        header_map={
            "signature": "Folder signature",
            "folder": "Folder",
            "open_folder": "Open folder link",
            "file_count": "File count",
            "total_bytes": "Total bytes",
            "total_size": "Total size",
            "matching_folder_count": "Matching folder count",
        },
    )
    write_table(wb, "Google Takeout Review", tables["Google Takeout Review"], FILE_COLUMNS)
    write_table(wb, "Large Files", tables["Large Files"], FILE_COLUMNS)
    write_table(wb, "Old Files", tables["Old Files"], FILE_COLUMNS)
    write_table(
        wb,
        "Extension Summary",
        tables["Extension Summary"],
        ["extension", "file_count", "total_bytes", "total_size"],
        header_map={
            "extension": "Extension",
            "file_count": "File count",
            "total_bytes": "Total bytes",
            "total_size": "Total size",
        },
    )
    write_table(
        wb,
        "Drive Summary",
        tables["Drive Summary"],
        ["drive", "file_count", "total_bytes", "total_size"],
        header_map={
            "drive": "Drive",
            "file_count": "File count",
            "total_bytes": "Total bytes",
            "total_size": "Total size",
        },
    )
    write_table(
        wb,
        "Folder Size Summary",
        tables["Folder Size Summary"][:max_rows],
        ["folder", "open_folder", "file_count", "total_bytes", "total_size"],
        header_map={
            "folder": "Folder",
            "open_folder": "Open folder link",
            "file_count": "File count",
            "total_bytes": "Total bytes",
            "total_size": "Total size",
        },
    )
    write_table(
        wb,
        "Scan Errors",
        tables["Scan Errors"],
        ["path", "phase", "error_type", "error_message", "created_at"],
        header_map={
            "path": "Path",
            "phase": "Phase",
            "error_type": "Error type",
            "error_message": "Error message",
            "created_at": "When",
        },
    )

    report_dir = Path(config["report_dir"])
    report_dir.mkdir(parents=True, exist_ok=True)
    if output_path is None:
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output = report_dir / f"storage_report_scan_{scan_id}_{stamp}.xlsx"
    else:
        output = Path(output_path)
    wb.save(output)
    return output


def export_csv_report(
    config_path: str | Path | None = None,
    output_dir: str | Path | None = None,
    scan_id: int | None = None,
    include_all_files: bool = False,
) -> Path:
    config = load_config(config_path)
    cache = StorageCache(config["database_path"])
    try:
        scan_id = scan_id or cache.latest_scan_id()
        if scan_id is None:
            raise RuntimeError("No scan found.")
        scan = cache.scan_row(scan_id)
        if scan is None:
            raise RuntimeError(f"Scan {scan_id} not found.")
        report_dir = Path(config["report_dir"])
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output = Path(output_dir) if output_dir else report_dir / f"storage_report_scan_{scan_id}_{stamp}_csv"
        output.mkdir(parents=True, exist_ok=True)

        metrics = basic_scan_metrics(scan, config)
        summary_rows = summary_rows_for(scan, metrics, cache.error_count(scan_id), config)
        write_csv(
            output / "summary.csv",
            ({"Metric": key, "Value": value} for key, value in summary_rows[1:]),
            ["Metric", "Value"],
        )
        write_csv(
            output / "duplicate_groups.csv",
            cache.iter_query_rows(
                """
                SELECT full_hash,
                       MIN(filename) AS sample_filename,
                       MIN(size) AS file_size_bytes,
                       COUNT(*) AS file_count,
                       (COUNT(*) - 1) * MIN(size) AS wasted_bytes,
                       COUNT(*) * MIN(size) AS involved_bytes,
                       GROUP_CONCAT(DISTINCT drive) AS drives
                FROM files
                WHERE last_seen_scan_id = ? AND full_hash IS NOT NULL
                GROUP BY full_hash, size
                HAVING COUNT(*) > 1
                ORDER BY wasted_bytes DESC
                """,
                (scan_id,),
            ),
            ["full_hash", "sample_filename", "file_size_bytes", "file_count", "wasted_bytes", "involved_bytes", "drives"],
        )
        write_csv(
            output / "duplicate_files.csv",
            cache.iter_query_rows(
                """
                SELECT f.*
                FROM files f
                JOIN (
                    SELECT full_hash, size
                    FROM files
                    WHERE last_seen_scan_id = ? AND full_hash IS NOT NULL
                    GROUP BY full_hash, size
                    HAVING COUNT(*) > 1
                ) candidates ON candidates.full_hash = f.full_hash AND candidates.size = f.size
                WHERE f.last_seen_scan_id = ?
                ORDER BY f.path
                """,
                (scan_id, scan_id),
            ),
            ["path", "folder", "drive", "filename", "extension", "size", "modified_at", "partial_hash", "full_hash"],
        )
        write_csv(
            output / "scan_errors.csv",
            cache.iter_error_rows(scan_id),
            ["path", "phase", "error_type", "error_message", "created_at"],
        )
        write_csv(
            output / "folder_summary.csv",
            cache.iter_query_rows(
                """
                SELECT folder, COUNT(*) AS file_count, COALESCE(SUM(size), 0) AS total_bytes
                FROM files
                WHERE last_seen_scan_id = ?
                GROUP BY folder
                ORDER BY total_bytes DESC
                """,
                (scan_id,),
            ),
            ["folder", "file_count", "total_bytes"],
        )
        if include_all_files:
            write_csv(
                output / "all_files.csv",
                cache.iter_file_rows(scan_id),
                ["path", "folder", "drive", "filename", "extension", "size", "modified_at", "partial_hash", "full_hash"],
            )
        return output
    finally:
        cache.close()


def basic_scan_metrics(scan: dict[str, Any], config: dict[str, Any]) -> dict[str, Any]:
    return {
        "total_files": int(scan.get("files_scanned") or 0),
        "total_size": int(scan.get("bytes_scanned") or 0),
        "duplicate_groups": int(scan.get("duplicate_groups") or 0),
        "duplicate_review_bytes": int(scan.get("duplicate_review_bytes") or 0),
        "duplicate_involved_bytes": int(scan.get("duplicate_involved_bytes") or 0),
        "biggest_duplicate_group_bytes": int(scan.get("biggest_duplicate_group_bytes") or 0),
        "files_already_backed_up_on_b": 0,
        "files_on_p_also_found_on_b": 0,
        "large_file_threshold_bytes": int(float(config.get("large_file_threshold_mb", 500)) * 1024 * 1024),
        "old_file_days": int(config.get("old_file_days", 365)),
    }


def export_scan_errors_csv(
    config_path: str | Path | None = None,
    output_path: str | Path | None = None,
    scan_id: int | None = None,
) -> Path:
    config = load_config(config_path)
    cache = StorageCache(config["database_path"])
    try:
        scan_id = scan_id or cache.latest_scan_id()
        if scan_id is None:
            raise RuntimeError("No scan found.")
        report_dir = Path(config["report_dir"])
        report_dir.mkdir(parents=True, exist_ok=True)
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output = Path(output_path) if output_path else report_dir / f"scan_errors_{scan_id}_{stamp}.csv"
        write_csv(
            output,
            cache.iter_error_rows(scan_id),
            ["path", "phase", "error_type", "error_message", "created_at"],
        )
        return output
    finally:
        cache.close()


def write_csv(path: Path, rows: Iterable[dict[str, Any]], columns: list[str]) -> None:
    with path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=columns, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def safe_file_name(name: str) -> str:
    cleaned = "".join(char if char.isalnum() else "_" for char in name.lower())
    return "_".join(part for part in cleaned.split("_") if part)


def summary_rows_for(
    scan: dict[str, Any],
    metrics: dict[str, Any],
    error_count: int,
    config: dict[str, Any],
) -> list[tuple[str, Any]]:
    return [
        ("Metric", "Value"),
        ("Scan ID", scan.get("id")),
        ("Mode", scan.get("mode")),
        ("Started", scan.get("started_at")),
        ("Completed", scan.get("completed_at")),
        ("Total files scanned", metrics["total_files"]),
        ("Total scanned size", human_bytes(metrics["total_size"])),
        ("Total scanned bytes", metrics["total_size"]),
        ("Duplicate groups", metrics["duplicate_groups"]),
        ("Duplicate size to review", human_bytes(metrics["duplicate_review_bytes"])),
        ("Duplicate bytes to review", metrics["duplicate_review_bytes"]),
        ("Duplicate space involved", human_bytes(metrics["duplicate_involved_bytes"])),
        ("Biggest duplicate group", human_bytes(metrics["biggest_duplicate_group_bytes"])),
        ("Files already backed up on B", metrics["files_already_backed_up_on_b"]),
        ("Files on P also found on B", metrics["files_on_p_also_found_on_b"]),
        ("Large file threshold", human_bytes(metrics["large_file_threshold_bytes"])),
        ("Old file cutoff days", metrics["old_file_days"]),
        ("Scan errors / permission denied", error_count),
        ("Safety", "Read-only. No delete, move, rename, or metadata change."),
        ("Review wording", "Safe to review manually"),
        ("Database", config["database_path"]),
    ]


def write_summary(wb: Workbook, scan: dict[str, Any], metrics: dict[str, Any], error_count: int, config: dict[str, Any]) -> None:
    ws = wb.create_sheet("Summary")
    rows = summary_rows_for(scan, metrics, error_count, config)
    for row in rows:
        ws.append(row)
    ws["A1"].font = Font(bold=True, color="FFFFFF")
    ws["B1"].font = Font(bold=True, color="FFFFFF")
    fill = PatternFill("solid", fgColor="1F2933")
    ws["A1"].fill = fill
    ws["B1"].fill = fill
    ws["B20"].comment = Comment("Use review label only. No destructive action here.", "Codex")
    ws.freeze_panes = "A2"
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 80
    for cell in ws["B"]:
        cell.alignment = Alignment(wrap_text=True, vertical="top")


def write_table(
    wb: Workbook,
    title: str,
    rows: list[dict[str, Any]],
    columns: list[str],
    header_map: dict[str, str] | None = None,
) -> None:
    header_map = header_map or HEADERS
    ws = wb.create_sheet(safe_sheet_name(title))
    headers = [header_map.get(col, col) for col in columns]
    ws.append(headers)
    for row in rows:
        ws.append([row.get(col, "") for col in columns])
    style_header(ws)
    ws.freeze_panes = "A2"
    if ws.max_row > 1:
        ws.auto_filter.ref = ws.dimensions
    set_widths(ws)
    for col_idx, col_name in enumerate(columns, start=1):
        if col_name in {"size", "total_bytes", "file_count", "age_days"}:
            for cell in ws.iter_cols(min_col=col_idx, max_col=col_idx, min_row=2):
                for item in cell:
                    item.number_format = "#,##0"
        if col_name in {"path", "folder", "labels", "error_message"}:
            for cell in ws.iter_cols(min_col=col_idx, max_col=col_idx, min_row=2):
                for item in cell:
                    item.alignment = Alignment(wrap_text=True, vertical="top")


def style_header(ws) -> None:
    fill = PatternFill("solid", fgColor="1F2933")
    font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(wrap_text=True, vertical="center")


def set_widths(ws) -> None:
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        max_len = 12
        for row in range(1, min(ws.max_row, 250) + 1):
            value = ws.cell(row=row, column=col).value
            if value is not None:
                max_len = max(max_len, min(len(str(value)) + 2, 80))
        ws.column_dimensions[letter].width = max_len


def safe_sheet_name(name: str) -> str:
    for char in "[]:*?/\\":
        name = name.replace(char, " ")
    return name[:31]
